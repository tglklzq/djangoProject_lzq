# myapp/views.py
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
import base64
from myapp.forms import UpdateProfileForm, DeleteCourseForm, AddCourseForm, SearchCourseTypeForm, AddCourseTypeForm, \
    CourseTypeForm, SearchForm, RegistrationForm
from myapp.model import User, Course, CourseType, Registration
import openpyxl
from django.http import HttpResponse
import io
from openpyxl.drawing.image import Image as XLImage
from django.shortcuts import get_object_or_404, render, redirect
from .forms import EditCourseForm  # 请在这里导入你的编辑表单
# 登录注册界面
def login_view(request):
    error_message = None

    if request.method == 'POST':
        user_number = request.POST.get('user_number')
        user_password = request.POST.get('user_password')

        try:
            user = User.objects.get(user_number=user_number)

            if user_password == user.user_password:
                request.session['user_number'] = user.user_number
                request.session['user_name'] = user.user_name
                return render(request, 'success.html', {'user': user})
            else:
                error_message = "账号或密码错误"
        except User.DoesNotExist:
            error_message = "用户不存在"

    return render(request, 'login.html', {'error_message': error_message})


def success_view(request):
    # 登录成功后的视图
    user = request.GET.get('user')
    return render(request, 'success.html', {'user': user})


# views.py


def register_view(request):
    if request.method == 'POST':
        user_number = request.POST.get('user_number')
        user_name = request.POST.get('user_name')
        user_password = request.POST.get('user_password')
        confirm_password = request.POST.get('confirm_password')

        # 检查密码是否匹配
        if user_password != confirm_password:
            return render(request, 'register.html', {'error_message': '密码不匹配'})

        # 创建用户
        user = User(user_number=user_number, user_name=user_name, user_password=user_password)
        user.save()

        # 注册成功后重定向到登录页面
        return redirect('login')

    return render(request, 'register.html')


def page_view(request, page_number):
    template_name = f'page/page{page_number}.html'
    return render(request, template_name)



def page5_view(request, user_number):
    user = User.objects.get(user_number=user_number)
    print("User Information:", user.user_name, user.user_number)  # 确保打印的用户信息正确
    return render(request, 'page/page5.html', {'user': user})


def update_profile_view(request):
    if request.method == 'POST':
        form = UpdateProfileForm(request.POST)
        if form.is_valid():
            new_name = form.cleaned_data['name']
            new_number = form.cleaned_data['number']
            new_password = form.cleaned_data['password']

            user_number = request.session.get('user_number')

            if user_number:
                User.objects.filter(user_number=user_number).update(
                    user_name=new_name,
                    user_number=new_number,
                    user_password=new_password
                )

                # 更新会话中的信息
                request.session['user_name'] = new_name
                request.session['user_number'] = new_number
                request.session.modified = True

                return redirect('login')

    else:
        form = UpdateProfileForm()

    return render(request, 'login.html', {'form': form})


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from .forms import AddCourseForm, DeleteCourseForm





# /////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# 以下是开发page1
def page1_view(request):
    courses = Course.objects.all()

    # 添加分页功能
    page = request.GET.get('page', 1)
    paginator = Paginator(courses, 20)

    try:
        courses = paginator.page(page)
    except PageNotAnInteger:
        courses = paginator.page(1)
    except EmptyPage:
        courses = paginator.page(paginator.num_pages)

    if request.method == 'POST':
        add_course_form = AddCourseForm(request.POST, request.FILES)  # 注意这里加入 request.FILES
        delete_course_form = DeleteCourseForm(request.POST)

        if add_course_form.is_valid():
            # 处理文件上传
            course_textbook_pic = request.FILES.get('course_textbook_pic')
            if course_textbook_pic:
                # 以二进制形式读取文件数据并存储到数据库
                add_course_form.instance.course_textbook_pic = course_textbook_pic.read()
            add_course_form.save()
            return redirect('page1')

        if delete_course_form.is_valid():
            course_no = delete_course_form.cleaned_data['course_no']
            Course.objects.filter(course_no=course_no).delete()
            return redirect('page1')
    else:
        add_course_form = AddCourseForm()
        delete_course_form = DeleteCourseForm()

    # 对每个课程的图片进行base64编码
    for course in courses:
        if course.course_textbook_pic:
            encoded_image = base64.b64encode(course.course_textbook_pic).decode('utf-8')
            course.encoded_image = encoded_image

    return render(request, 'page/page1.html',
                  {'courses': courses, 'add_course_form': add_course_form, 'delete_course_form': delete_course_form})





def export_to_excel(request):
    # 从数据库中获取课程数据
    all_courses = Course.objects.all()

    # 创建一个新的Excel文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 添加表头
    sheet.append(['课程编号', '课程名称', '课时数', '课程类型', '课程状态',
                  '要求', '学分', '备注', '书本封面'])

    # 遍历课程数据，并将其写入到Excel文件中
    for course in all_courses:
        sheet.append([course.course_no, course.course_name, course.course_hours, course.type_id,
                      course.course_status, course.course_reqs, course.course_point, course.course_memo])

        # 插入图片到Excel
        if course.course_textbook_pic:
            img = XLImage(io.BytesIO(course.course_textbook_pic))
            img.width = img.width * 0.5  # 调整图片大小
            img.height = img.height * 0.5
            sheet.add_image(img, f'I{sheet.max_row}')  # 假设将图片插入到H列
            sheet.row_dimensions[sheet.max_row].height = img.height

    # 保存Excel文件
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=courses.xlsx'
    workbook.save(response)
    return response





def edit_course_view(request, course_no):
    course = get_object_or_404(Course, course_no=course_no)

    if request.method == 'POST':
        edit_course_form = EditCourseForm(request.POST, request.FILES, instance=course)
        if edit_course_form.is_valid():
            edit_course_form.save()
            return redirect('page1')
    else:
        edit_course_form = EditCourseForm(instance=course)

    return render(request, 'page/edit_course.html', {'edit_course_form': edit_course_form})


def delete_course_view(request, course_no):
    course = Course.objects.get(course_no=course_no)
    # 在这里添加删除课程的逻辑
    course.delete()
    return redirect('page1')

# page1查询功能：
def search_courses_type(request):
    if 'search_query' in request.GET:
        search_query = request.GET['search_query']
        # 在这里执行你的搜索逻辑，例如通过模型查询数据库
        # 使用 contains 进行模糊匹配，这里以课程名称为例
        search_results = Course.objects.filter(course_name__contains=search_query)

        # 对每个查询结果的图片进行base64编码
        for result in search_results:
            if result.course_textbook_pic:
                encoded_image = base64.b64encode(result.course_textbook_pic).decode('utf-8')
                result.encoded_image = encoded_image

        return render(request, 'page/search_results.html',
                      {'search_results': search_results, 'search_query': search_query})
    else:
        # 如果没有搜索查询，则重定向到原始页面或执行其他操作
        return render(request, 'page/page1.html', {'courses': Course.objects.all()})


# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
# 开发page2：


from django.shortcuts import render, redirect

from .forms import CourseTypeForm


def page2_view(request):
    # 获取所有课程类型
    course_types = CourseType.objects.all()

    # 渲染模板
    return render(request, 'page/page2.html', {'course_types': course_types})


def add_course_type_view(request):
    if request.method == 'POST':
        form = CourseTypeForm(request.POST)
        if form.is_valid():
            type_id = form.cleaned_data['type_id']
            type_name = form.cleaned_data['type_name']

            # 创建CourseType实例并保存到数据库
            new_course_type = CourseType(type_id=type_id, type_name=type_name)
            new_course_type.save()

            return redirect('page2')
    else:
        form = CourseTypeForm()

    return render(request, 'page/add_course_type.html', {'form': form})


def edit_course_type_view(request, type_id):
    # 获取要编辑的课程类型对象
    course_type = get_object_or_404(CourseType, pk=type_id)

    if request.method == 'POST':
        # 处理表单提交
        form = CourseTypeForm(request.POST)
        if form.is_valid():
            # 手动更新表单关联的模型实例
            course_type.type_name = form.cleaned_data['type_name']
            course_type.save()
            print("Course type updated successfully!")  # 添加调试输出
            return redirect('page2')
        else:
            print("Form validation failed:", form.errors)  # 添加调试输出
    else:
        # 显示包含当前对象数据的表单
        form = CourseTypeForm(initial={'type_name': course_type.type_name})

    return render(request, 'page/edit_course_type.html', {'form': form, 'course_type': course_type})


def delete_course_type_view(request, type_id):
    # 获取要删除的课程类型对象
    course_type = CourseType.objects.get(pk=type_id)

    if request.method == 'POST':
        # 处理删除请求
        course_type.delete()
        return redirect('page2')

    return render(request, 'page/delete_course_type.html', {'course_type': course_type})


def search_course_type_view(request):
    query = request.GET.get('search_query', '')

    if query:
        # 如果查询参数不为空，则尝试通过 ID 或类型名称进行查询
        course_types = CourseType.objects.filter(type_id__icontains=query) | CourseType.objects.filter(
            type_name__icontains=query)
        return render(request, 'page/search_course_type.html', {'course_types': course_types, 'query': query})
    else:
        # 如果查询参数为空，则显示所有信息
        course_types = CourseType.objects.all()
        return render(request, 'page/search_course_type.html', {'course_types': course_types, 'query': query})


# 以下进行开发page4.html///////////////////


def page4_view(request):
    registrations = Registration.objects.all()
    search_results = []

    if request.GET.get('search'):
        query = request.GET.get('search')
        search_results = Registration.objects.filter(teacher__icontains=query) | Registration.objects.filter(
            course__icontains=query)

    context = {
        'registrations': registrations,
        'search_results': search_results
    }

    return render(request, 'page/page4.html', context)


def add_registration(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('page4')  # 这里的'page4'是你页面4的URL名称
    else:
        form = RegistrationForm()

    return render(request, 'page/page4.html', {'form': form})


def edit_registration_view(request, id):
    # 获取要编辑的记录对象
    registration = Registration.objects.get(pk=id)

    if request.method == 'POST':
        # 处理编辑记录的逻辑
        form = RegistrationForm(request.POST, instance=registration)
        if form.is_valid():
            form.save()
            return redirect('page4')
    else:
        form = RegistrationForm(instance=registration)

    return render(request, 'page/edit_registration.html', {'form': form, 'registration': registration})


def delete_registration_view(request, id):
    # 获取要删除的记录对象
    registration = Registration.objects.get(pk=id)

    if request.method == 'POST':
        # 处理删除记录的逻辑
        registration.delete()
        return redirect('page4')

    return render(request, 'page/delete_registration.html', {'registration': registration})
